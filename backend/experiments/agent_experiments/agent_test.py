"""
NutriRAG Agent Evaluation (Snowflake Cortex)
- Runs the Cortex Agent (SSE stream) and collects:
  - agent_answer, agent_thinking, tool_results, raw events
- Evaluates with:
  (A) DeepEval metrics (optional): Faithfulness, AnswerRelevancy
  (B) Custom LLM-as-judge (Cortex inference:complete): tool_dependency, tool_selection, tool_sequence, tool_params, correctness
- Saves everything into an Excel file (.xlsx)

Requirements:
  pip install requests openpyxl
Optional (if you want DeepEval metrics):
  pip install deepeval

Assumes you already have:
  from shared.snowflake.client import SnowflakeClient
"""

import os
import sys
import json
import time
import re
import requests
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))
from shared.snowflake.client import SnowflakeClient


# =============================================================================
# 1) SSE parsing utilities
# =============================================================================

def iter_sse_events(response: requests.Response):
    """
    Minimal SSE parser.
    Yields dicts: {"event": str|None, "data": str|None, "raw": [lines]}
    """
    event_name = None
    data_lines: List[str] = []
    raw_lines: List[str] = []

    for line in response.iter_lines(decode_unicode=True):
        if line is None:
            continue

        raw_lines.append(line)

        if line == "":
            if event_name is not None or data_lines:
                yield {
                    "event": event_name,
                    "data": "\n".join(data_lines) if data_lines else None,
                    "raw": raw_lines,
                }
            event_name = None
            data_lines = []
            raw_lines = []
            continue

        if line.startswith(":"):
            # keep-alive/comments
            continue

        if line.startswith("event:"):
            event_name = line[len("event:"):].strip()
        elif line.startswith("data:"):
            data_lines.append(line[len("data:"):].strip())

    if event_name is not None or data_lines:
        yield {
            "event": event_name,
            "data": "\n".join(data_lines) if data_lines else None,
            "raw": raw_lines,
        }


def safe_json_loads(s: Optional[str]) -> Optional[Any]:
    if not s:
        return None
    try:
        return json.loads(s)
    except Exception:
        return None


def extract_first_json_object(text: str) -> Optional[dict]:
    """
    Robust parser for judge outputs that may contain JSON + extra explanation.
    We extract the FIRST {...} block and try json.loads on it.
    """
    if not text:
        return None
    m = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if not m:
        return None
    candidate = m.group(0).strip()
    try:
        return json.loads(candidate)
    except Exception:
        return None


# =============================================================================
# 2) Agent client (Cortex Agents API)
# =============================================================================

class CortexAgentClient:
    def __init__(self, snowflake_client: SnowflakeClient):
        self.sf = snowflake_client

    def call_agent_stream(self, prompt: str, include_thinking: bool = True) -> requests.Response:
        token = self.sf.get_jwt()
        cfg = self.sf.config

        account = cfg["account"]
        db = cfg["database"]
        schema = cfg["schema_agent"]
        agent = cfg["agent"]

        url = f"https://{account}.snowflakecomputing.com/api/v2/databases/{db}/schemas/{schema}/agents/{agent}:run"

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "Accept": "application/json, text/event-stream",
            "X-Snowflake-Authorization-Token-Type": "KEYPAIR_JWT",
        }

        body = {
            "messages": [{"role": "user", "content": [{"type": "text", "text": prompt}]}],
            "tool_choice": {"type": "auto"},
            "include_thinking": include_thinking,
        }

        return requests.post(url, headers=headers, json=body, stream=True, timeout=180)


def run_agent_and_collect(
    agent_client: CortexAgentClient,
    prompt: str,
    save_path: str,
    include_thinking: bool = True,
    print_thinking: bool = False,
    print_raw: bool = False,
) -> Tuple[str, str, List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Returns:
      agent_answer, agent_thinking, tool_results, all_events
    """
    resp = agent_client.call_agent_stream(prompt, include_thinking=include_thinking)
    resp.raise_for_status()

    all_events: List[Dict[str, Any]] = []
    tool_results: List[Dict[str, Any]] = []

    final_text_parts: List[str] = []
    final_thinking_parts: List[str] = []

    for ev in iter_sse_events(resp):
        all_events.append(ev)

        if print_raw:
            for raw in ev.get("raw", []):
                if raw.strip():
                    print(raw)

        event_name = ev.get("event")
        payload = safe_json_loads(ev.get("data"))

        if event_name == "response.thinking.delta" and isinstance(payload, dict):
            txt = payload.get("text", "")
            if txt:
                final_thinking_parts.append(txt)
                if print_thinking:
                    print(txt, end="", flush=True)

        if event_name == "response.text.delta" and isinstance(payload, dict):
            txt = payload.get("text", "")
            if txt:
                final_text_parts.append(txt)

        if event_name == "response.tool_result" and isinstance(payload, dict):
            tool_results.append(payload)

    agent_answer = "".join(final_text_parts).strip()
    agent_thinking = "".join(final_thinking_parts).strip()

    with open(save_path, "w", encoding="utf-8") as f:
        json.dump(all_events, f, ensure_ascii=False, indent=2)

    return agent_answer, agent_thinking, tool_results, all_events


def extract_tool_outputs_pretty(tool_results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Convert tool_result payloads into a simpler dict:
      {
        "search": [<tool json>...],
        "transform": [...],
      }
    """
    out: Dict[str, Any] = {}

    for tr in tool_results:
        tool_name = tr.get("name") or tr.get("tool_name") or "unknown_tool"
        content = tr.get("content")

        if isinstance(content, list):
            extracted = []
            for item in content:
                if isinstance(item, dict) and "json" in item:
                    extracted.append(item["json"])
                else:
                    extracted.append(item)
            out.setdefault(tool_name, []).extend(extracted)
        else:
            out.setdefault(tool_name, []).append(tr)

    return out


def extract_tools_called_from_events(all_events: List[Dict[str, Any]]) -> List[str]:
    """
    Reads SSE events to extract the *order* of tool calls from "response.tool_use".
    This is better than using tool_outputs keys (which loses ordering).
    """
    tools_in_order: List[str] = []
    for ev in all_events:
        if ev.get("event") != "response.tool_use":
            continue
        payload = safe_json_loads(ev.get("data"))
        if isinstance(payload, dict):
            name = payload.get("name")
            if name:
                tools_in_order.append(name)
    return tools_in_order


# =============================================================================
# 3) Judge client (Cortex inference:complete) - non streaming
# =============================================================================

class CortexJudgeClient:
    def __init__(self, snowflake_client: SnowflakeClient, model: str = "claude-3-5-sonnet"):
        self.sf = snowflake_client
        self.model = model

    def judge(self, judge_prompt: str) -> Dict[str, Any]:
        token = self.sf.get_jwt()
        account = self.sf.config["account"]
        url = f"https://{account}.snowflakecomputing.com/api/v2/cortex/inference:complete"

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "X-Snowflake-Authorization-Token-Type": "KEYPAIR_JWT",
        }

        body = {
            "model": self.model,
            "messages": [{"role": "user", "content": judge_prompt}],
            "max_tokens": 1400,
            "temperature": 0,
            "top_p": 1,
            "stream": False,
        }

        r = requests.post(url, headers=headers, json=body, timeout=180)
        r.raise_for_status()
        return r.json()


# =============================================================================
# 4) Custom LLM-as-judge prompt (NO clarity)
# =============================================================================

def build_custom_judge_prompt(
    user_prompt: str,
    agent_answer: str,
    tool_outputs: Dict[str, Any],
    tools_called_in_order: List[str],
) -> str:
    """
    Outputs JSON only.
    Scores are 0-5.
    """
    return f"""
You are an evaluator ("LLM-as-judge") for a tool-using cooking assistant agent called NutriRAG.

You MUST output a single JSON object (no markdown, no extra text) with this exact shape:
{{
  "scores": {{
    "tool_dependency": <0-5>,
    "tool_selection": <0-5>,
    "tool_sequence": <0-5>,
    "tool_params": <0-5>,
    "correctness": <0-5>
  }},
  "pass": <true/false>,
  "issues": [<string>, ...],
  "suggested_fix": <string>
}}

Rubric:
- tool_dependency:
  If tools fail/are empty, does the agent avoid improvising and guide next steps?
  5 = perfectly handles missing tools without fabricating.
- tool_selection:
  Did it call the right tools for the intent (search vs transform)? No missing critical tool calls.
- tool_sequence:
  Did it call tools in a logically correct order (search before transform, etc.)?
- tool_params:
  Are filters/constraints correctly extracted from the user's request and passed/applied consistently?
  Penalize missing constraints, wrong operators, or inventing filters.
- correctness:
  Given tool outputs and user request, is the response logically correct (no contradictions)?

Key rule:
The agent must NOT invent recipes/ingredients/steps/times/nutrition/IDs or claims not found in tool outputs.

User prompt:
{user_prompt}

Tools called (in order, from SSE):
{json.dumps(tools_called_in_order, ensure_ascii=False)}

Agent answer:
{agent_answer if agent_answer else "(empty)"}

Tool outputs (source of truth):
{json.dumps(tool_outputs, ensure_ascii=False, indent=2)}

Now produce the JSON verdict ONLY.
""".strip()


# =============================================================================
# 5) DeepEval (optional) + Cortex evaluator wrapper (streaming)
# =============================================================================

DEEPEVAL_AVAILABLE = True
try:
    from deepeval.models.base_model import DeepEvalBaseLLM
    from deepeval.metrics import FaithfulnessMetric, AnswerRelevancyMetric
    from deepeval.test_case import LLMTestCase
except Exception:
    DEEPEVAL_AVAILABLE = False


class SnowflakeCortexDeepEvalLLM(DeepEvalBaseLLM):  # type: ignore
    """
    DeepEval LLM wrapper using Cortex inference:complete (SSE streaming).
    DeepEval expects `generate(prompt)->str`.
    """

    def __init__(self, snowflake_client: SnowflakeClient, model: str = "claude-3-5-sonnet"):
        self.sf = snowflake_client
        self.model = model

    def load_model(self):
        return self.model

    def generate(self, prompt: str) -> str:
        token = self.sf.get_jwt()
        account = self.sf.config["account"]
        url = f"https://{account}.snowflakecomputing.com/api/v2/cortex/inference:complete"

        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "Accept": "text/event-stream",
            "X-Snowflake-Authorization-Token-Type": "KEYPAIR_JWT",
        }

        body = {
            "model": self.model,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": 1200,
            "temperature": 0,
            "stream": True,
        }

        r = requests.post(url, headers=headers, json=body, stream=True, timeout=180)
        if r.status_code != 200:
            return ""

        chunks: List[str] = []
        for line in r.iter_lines(decode_unicode=True):
            if not line:
                continue
            if not line.startswith("data: "):
                continue
            data_str = line[6:].strip()
            if data_str == "[DONE]":
                break
            try:
                data_json = json.loads(data_str)
                choices = data_json.get("choices", [])
                if choices:
                    delta = choices[0].get("delta", {})
                    content = delta.get("content", "")
                    if content:
                        chunks.append(content)
            except Exception:
                continue

        text = "".join(chunks).strip()
        # remove ```json fences if any
        if text.startswith("```json"):
            text = text[7:].lstrip()
        if text.endswith("```"):
            text = text[:-3].rstrip()
        return text

    async def a_generate(self, prompt: str) -> str:
        return self.generate(prompt)

    def get_model_name(self):
        return self.model


def build_retrieval_context_for_deepeval(tool_results: List[Dict[str, Any]]) -> List[str]:
    """
    DeepEval uses retrieval_context for Faithfulness / Relevancy grounding.
    We'll store a compact string per tool_result.
    """
    context_list: List[str] = []
    for tr in tool_results:
        tool_name = tr.get("name") or tr.get("tool_name") or "unknown"
        content = tr.get("content")
        clean = content
        if isinstance(content, list):
            clean = []
            for item in content:
                if isinstance(item, dict) and "json" in item:
                    clean.append(item["json"])
                else:
                    clean.append(item)
        context_list.append(f"Tool '{tool_name}' Output: {json.dumps(clean, ensure_ascii=False)}")
    return context_list


# =============================================================================
# 6) Excel writer
# =============================================================================

EXCEL_COLUMNS = [
    "timestamp",
    "user_prompt",
    "agent_answer",
    "tools_called_order",
    "execution_time_ms",

    # custom judge
    "custom_pass",
    "custom_tool_dependency",
    "custom_tool_selection",
    "custom_tool_sequence",
    "custom_tool_params",
    "custom_correctness",
    "custom_issues",
    "custom_suggested_fix",

    # DeepEval (optional)
    "deepeval_faithfulness",
    "deepeval_faithfulness_reason",
    "deepeval_answer_relevancy",
    "deepeval_answer_relevancy_reason",

    # debug blobs (optional but useful)
    "tool_outputs_json",
    "events_path",
]


def autosize_worksheet(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)


def append_row_to_excel(xlsx_path: str, row: Dict[str, Any]):
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "eval"
        ws.append(EXCEL_COLUMNS)

    values = [row.get(col, "") for col in EXCEL_COLUMNS]
    ws.append(values)

    autosize_worksheet(ws)
    wb.save(xlsx_path)


# =============================================================================
# 7) One evaluation run
# =============================================================================

def evaluate_one_prompt(
    sf_client: SnowflakeClient,
    user_prompt: str,
    excel_path: str = "nutrirag_eval.xlsx",
    judge_model: str = "claude-3-5-sonnet",
    deepeval_model: str = "claude-3-5-sonnet",
    save_dir: str = "eval_runs",
    print_thinking: bool = False,
) -> Dict[str, Any]:
    os.makedirs(save_dir, exist_ok=True)

    agent_client = CortexAgentClient(sf_client)
    judge_client = CortexJudgeClient(sf_client, model=judge_model)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    events_path = os.path.join(save_dir, f"agent_events_{ts}.json")

    # ---- Run agent
    start = time.time()
    agent_answer, agent_thinking, tool_results, all_events = run_agent_and_collect(
        agent_client=agent_client,
        prompt=user_prompt,
        save_path=events_path,
        include_thinking=True,
        print_thinking=print_thinking,
        print_raw=False,
    )
    execution_time_ms = (time.time() - start) * 1000.0

    tool_outputs = extract_tool_outputs_pretty(tool_results)
    tools_called_order = extract_tools_called_from_events(all_events)

    # ---- Custom judge
    custom_prompt = build_custom_judge_prompt(
        user_prompt=user_prompt,
        agent_answer=agent_answer,
        tool_outputs=tool_outputs,
        tools_called_in_order=tools_called_order,
    )
    judge_json = judge_client.judge(custom_prompt)
    judge_text = ""
    try:
        judge_text = judge_json["choices"][0]["message"]["content"]
    except Exception:
        judge_text = ""

    custom_verdict = extract_first_json_object(judge_text) or {}
    custom_scores = custom_verdict.get("scores", {}) if isinstance(custom_verdict, dict) else {}

    # ---- DeepEval (optional)
    deepeval_faithfulness = ""
    deepeval_faithfulness_reason = ""
    deepeval_answer_rel = ""
    deepeval_answer_rel_reason = ""

    if DEEPEVAL_AVAILABLE:
        try:
            cortex_deepeval_llm = SnowflakeCortexDeepEvalLLM(sf_client, model=deepeval_model)
            retrieval_context = build_retrieval_context_for_deepeval(tool_results)

            test_case = LLMTestCase(
                input=user_prompt,
                actual_output=agent_answer,
                retrieval_context=retrieval_context,
            )

            faith = FaithfulnessMetric(
                threshold=0.7,
                model=cortex_deepeval_llm,
                include_reason=True,
            )
            rel = AnswerRelevancyMetric(
                threshold=0.7,
                model=cortex_deepeval_llm,
                include_reason=True,
            )

            faith.measure(test_case)
            rel.measure(test_case)

            deepeval_faithfulness = str(faith.score)
            deepeval_faithfulness_reason = str(faith.reason or "")
            deepeval_answer_rel = str(rel.score)
            deepeval_answer_rel_reason = str(rel.reason or "")

        except Exception as e:
            deepeval_faithfulness = "ERR"
            deepeval_faithfulness_reason = f"{type(e).__name__}: {e}"
            deepeval_answer_rel = "ERR"
            deepeval_answer_rel_reason = f"{type(e).__name__}: {e}"

    # ---- Row to Excel
    row = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "user_prompt": user_prompt,
        "agent_answer": agent_answer,
        "tools_called_order": json.dumps(tools_called_order, ensure_ascii=False),
        "execution_time_ms": round(execution_time_ms, 1),

        "custom_pass": str(custom_verdict.get("pass", "")),
        "custom_tool_dependency": custom_scores.get("tool_dependency", ""),
        "custom_tool_selection": custom_scores.get("tool_selection", ""),
        "custom_tool_sequence": custom_scores.get("tool_sequence", ""),
        "custom_tool_params": custom_scores.get("tool_params", ""),
        "custom_correctness": custom_scores.get("correctness", ""),
        "custom_issues": json.dumps(custom_verdict.get("issues", []), ensure_ascii=False),
        "custom_suggested_fix": custom_verdict.get("suggested_fix", ""),

        "deepeval_faithfulness": deepeval_faithfulness,
        "deepeval_faithfulness_reason": deepeval_faithfulness_reason,
        "deepeval_answer_relevancy": deepeval_answer_rel,
        "deepeval_answer_relevancy_reason": deepeval_answer_rel_reason,

        "tool_outputs_json": json.dumps(tool_outputs, ensure_ascii=False),
        "events_path": events_path,
    }

    append_row_to_excel(excel_path, row)
    return row


# =============================================================================
# 8) Main (edit prompts here)
# =============================================================================

if __name__ == "__main__":
    sf_client = SnowflakeClient()

    # prompts = [
    #     # Simple baseline
    #     "Find me a recipe with chicken under 30 minutes",

    #     # Complex (search + transform)
    #     "Je veux 8 recettes avec du poulet (obligatoire) prêtes en <= 25 minutes. "
    #     "Interdits: lait, fromage, crème, beurre, noix/amandes/cacahuètes. "
    #     "Je veux AU MOINS une option low_carb et une option low_sodium si possible. "
    #     "Ensuite, transforme la recette #2 pour réduire les calories et le sodium (sans ajouter de nouveaux ingrédients) "
    #     "et dis-moi le before/after.",
    # ]
    # prompts = [
    # # ----------------------------
    # # A) Baseline / sanity checks
    # # ----------------------------
    # "chocolate cake",
    # "pasta with tomato sauce",
    # "healthy breakfast ideas",
    # "quick dinner ideas",

    # # -----------------------------------------
    # # B) Français + bruit (bonjour/merci etc.)
    # # -----------------------------------------
    # "Bonjour ! Je cherche une recette de pâtes.",
    # "Salut, merci ! Une idée de petit-déjeuner sain ?",
    # "Coucou, je veux un dessert au citron.",
    # "Bonsoir, je veux un repas rapide à faire ce soir.",

    # # --------------------------------------------
    # # C) Temps / contraintes numériques (minutes)
    # # --------------------------------------------
    # "Find me 5 recipes under 20 minutes",
    # "Je veux 5 recettes prêtes en moins de 30 minutes",
    # "Vegetarian dinner under 25 minutes",
    # "Dîner vegan en <= 20 minutes",

    # # ----------------------------
    # # D) Tags alimentaires
    # # ----------------------------
    # "Give me 5 vegan desserts",
    # "Gluten-free breakfast ideas",
    # "Dairy-free dinner recipes",
    # "Low carb lunch recipes",
    # "Diabetic-friendly meal ideas",

    # # ---------------------------------------
    # # E) Ingrédients include / exclude / any
    # # ---------------------------------------
    # "Recipes with chicken and rice",
    # "Pasta with tomato and basil",
    # "Dessert without nuts",
    # "No dairy and no eggs dessert",
    # "Dinner with either tofu or chickpeas",

    # # -----------------------------------------
    # # F) Conceptuel (vector/hybrid)
    # # -----------------------------------------
    # "light and refreshing summer dinner",
    # "comfort food but not too heavy",
    # "high-protein post-workout meal ideas",

    # # -----------------------------------------
    # # G) Cas pièges / robustesse
    # # -----------------------------------------
    # "bonjour salut merci",  # devrait déclencher une demande de clarification côté agent
    # "vegan but with chicken",  # contradiction à gérer proprement
    # "gluten-free bread with wheat flour",  # contradiction
    # "I want something healthy",  # ambigu -> ton agent doit demander 1 question
    # "gourmet meal under 10 minutes",  # potentiellement 0 résultats

    # # -----------------------------------------
    # # H) Stress test (k élevé, filtres multiples)
    # # -----------------------------------------
    # "Je veux 10 recettes avec poulet ET riz, prêtes en <= 25 minutes, sans lactose, sans gluten.",
    # ]

    prompts = [
# =============================================================================
# 1) Multi-constraints très denses (temps + tags + ingrédients + exclusions)
# =============================================================================
"Je veux 7 recettes de dîner: poulet obligatoire, < 25 minutes, sans lactose et sans gluten. "
"Interdits: noix, amandes, cacahuètes, crème, beurre. "
"Doit contenir AU MOINS un de ces ingrédients: citron OU gingembre OU paprika. "
"Et pas de champignons ni d’olives.",

"Give me 12 recipes with either salmon OR tofu, but NO shellfish, under 30 minutes, "
"and must be low_sodium. Exclude: butter, cream, cheese. Include: garlic AND lemon.",

"Je veux 4 recettes de petit-déjeuner vegan, <= 15 minutes, low_calorie si possible. "
"Je veux absolument: avoine OU banane (au moins un des deux). "
"Mais sans fruits à coque, sans lait, sans œufs.",

# =============================================================================
# 2) Tests opérateurs numériques (>= <= =) + objectifs nutrition
# =============================================================================
"Find 6 recipes with protein >= 30g and calories <= 450, under 35 minutes. "
"Must be gluten_free. Exclude: peanuts, almonds, walnuts.",

"Je veux 5 recettes avec carbs <= 25g et sodium <= 600mg, sans lactose, "
"et prêtes en moins de 40 minutes. Ingrédients obligatoires: tomate ET ail.",

"Give me 3 diabetic-friendly dinners with calories < 500 and fiber >= 8g. "
"Include chicken OR turkey, exclude rice and pasta.",

# =============================================================================
# 3) Include vs Any vs Exclude (logique booléenne)
# =============================================================================
"Je veux 8 recettes qui contiennent: (poulet ET riz) "
"ET aussi (au moins un de: poivron OU courgette OU aubergine). "
"Sans: soja, miel, cacahuètes.",

"Find me 10 pasta recipes that MUST include basil AND tomato, "
"but must NOT include cheese OR cream. Any of: tuna OR chicken.",

"Dessert: je veux 6 recettes avec chocolat OU citron, "
"mais sans noix, sans œufs, sans lait, et <= 45 minutes.",

# =============================================================================
# 6) Bruit + langue mixte + contraintes cachées (traduction + nettoyage)
# =============================================================================
"Bonjour chef !! stp merci :) Je voudrais, genre, un truc pour dîner... "
"chicken + rice, pas trop gras, et vite fait (moins de 20 min). "
"Ah et no dairy, no nuts. 6 recettes.",

"Salut!! Je suis pressé: 'quick lunch' mais aussi 'low sodium' "
"et 'high protein'. Pas de poisson. 5 options.",

"Hey! I need ideas for 'comfort food' but low_calorie, under 25 minutes, "
"no_lactose. Also exclude: bacon, sausage, ham. Give me 8.",

# =============================================================================
# 7) Requête très longue type cahier des charges (stress maximal)
# =============================================================================
"Je prépare un menu pour 3 jours. Donne-moi 9 recettes au total: "
"3 petits-déjeuners + 3 déjeuners + 3 dîners. "
"Contraintes globales: sans lactose, sans gluten, nut_free. "
"Pour les dîners: poulet OU tofu obligatoire, <= 30 minutes. "
"Pour les déjeuners: low_sodium obligatoire. "
"Pour les petits-déj: low_calorie si possible. "
"Je déteste: champignons, olives, coriandre. "
"J’adore: citron, ail, tomates. "
"Important: garde exactement 9 recettes.",

# =============================================================================
# 8) Tests sur k_input (singulier/pluriel/nombre explicite)
# =============================================================================
"Donne-moi UNE recette avec poulet et riz, <= 25 minutes, sans lactose.",

"Give me exactly 11 vegetarian dinners under 30 minutes.",

# =============================================================================
# 9) Pièges de reformulation (ne pas inventer)
# =============================================================================
"Je veux une recette 'comme un butter chicken' mais sans beurre, sans crème, "
"et sans produits laitiers. 5 recettes. (ne rajoute pas d’autres contraintes)",

"Find recipes inspired by Japanese flavors (miso, ginger, soy) but exclude soy sauce. "
"Under 30 minutes. Return 7.",

# =============================================================================
# 10) Cas extrême: contraintes probablement impossibles → total_found=0
# =============================================================================
"Je veux 10 recettes vegan, sans gluten, sans lactose, sans légumes, "
"<= 10 minutes, et avec poulet obligatoire.",

"Give me 5 recipes that are simultaneously: low_carb, low_calorie, high_protein, "
"under 10 minutes, and must include beef AND be vegan.",
]


    excel_path = "nutrirag_eval.xlsx"

    for p in prompts:
        print("\n" + "=" * 80)
        print("PROMPT:", p)
        result_row = evaluate_one_prompt(
            sf_client=sf_client,
            user_prompt=p,
            excel_path=excel_path,
            judge_model="claude-3-5-sonnet",
            deepeval_model="claude-3-5-sonnet",
            save_dir="eval_runs",
            print_thinking=False,
        )
        print("Saved to Excel:", excel_path)
        print("Custom pass:", result_row["custom_pass"])
